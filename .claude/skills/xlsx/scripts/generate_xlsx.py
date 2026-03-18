#!/usr/bin/env python3
"""
XLSX Generator — Claude Code Skill
Usage: python3 generate_xlsx.py <data.json>
Requires: pip install openpyxl
"""

import json
import sys
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl -q")
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule

# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

TOTAL_FILL    = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FONT    = Font(name="Calibri", bold=True, size=11)

ALT_FILL      = PatternFill("solid", fgColor="EBF5FB")
NORMAL_FONT   = Font(name="Calibri", size=10)
NORMAL_ALIGN  = Alignment(vertical="center")

THIN_BORDER   = Border(
    left   = Side(style="thin", color="BDC3C7"),
    right  = Side(style="thin", color="BDC3C7"),
    top    = Side(style="thin", color="BDC3C7"),
    bottom = Side(style="thin", color="BDC3C7"),
)


def style_header_row(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row: int, end_row: int, col_count: int):
    for row in range(start_row, end_row + 1):
        fill = ALT_FILL if row % 2 == 0 else None
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font      = NORMAL_FONT
            cell.alignment = NORMAL_ALIGN
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill


def style_totals_row(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = TOTAL_FILL
        cell.font   = TOTAL_FONT
        cell.border = THIN_BORDER


def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 45)


def add_chart(ws, chart_cfg: dict, data_rows: int, headers: list):
    chart_type = chart_cfg.get("type", "bar").lower()
    title      = chart_cfg.get("title", "Chart")
    x_col      = chart_cfg.get("x_col", 0) + 1
    y_cols     = [c + 1 for c in chart_cfg.get("y_cols", [1])]
    anchor     = chart_cfg.get("anchor", "H2")

    if chart_type == "bar":
        chart = BarChart()
    elif chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    else:
        chart = BarChart()

    chart.title  = title
    chart.height = 12
    chart.width  = 20

    for y_col in y_cols:
        data = Reference(ws, min_col=y_col, min_row=1, max_row=data_rows + 1)
        chart.add_data(data, titles_from_data=True)

    cats = Reference(ws, min_col=x_col, min_row=2, max_row=data_rows + 1)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def process_sheet(wb: Workbook, sheet_cfg: dict):
    name    = sheet_cfg.get("name", "Sheet")
    headers = sheet_cfg.get("headers", [])
    rows    = sheet_cfg.get("rows", [])
    has_totals  = sheet_cfg.get("totals_row", False)
    freeze_hdr  = sheet_cfg.get("freeze_header", True)
    auto_filter = sheet_cfg.get("auto_filter", True)
    chart_cfg   = sheet_cfg.get("chart")

    ws = wb.create_sheet(title=name)

    # Write headers
    if headers:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        style_header_row(ws, 1, len(headers))

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    style_data_rows(ws, 2, len(rows) + 1, len(headers))

    # Totals row
    if has_totals and rows:
        totals_row = len(rows) + 2
        ws.cell(row=totals_row, column=1, value="TOTAL")
        for col_idx in range(2, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=totals_row, column=col_idx,
                    value=f"=SUM({col_letter}2:{col_letter}{totals_row - 1})")
        style_totals_row(ws, totals_row, len(headers))

    if freeze_hdr:
        ws.freeze_panes = "A2"
    if auto_filter and headers:
        ws.auto_filter.ref = ws.dimensions

    autofit_columns(ws)
    ws.row_dimensions[1].height = 30

    if chart_cfg:
        add_chart(ws, chart_cfg, len(rows), headers)

    return ws


def generate(data_path: str):
    with open(data_path) as f:
        config = json.load(f)

    output = config.get("filename", "output.xlsx")
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for sheet_cfg in config.get("sheets", []):
        process_sheet(wb, sheet_cfg)

    wb.save(output)
    sheet_names = [s.get("name") for s in config.get("sheets", [])]
    print(f"✅  Saved {len(sheet_names)} sheet(s): {', '.join(sheet_names)} → {output}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: generate_xlsx.py <data.json>")
        sys.exit(1)
    generate(sys.argv[1])
